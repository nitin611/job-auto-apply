"""Daily Excel export — 4 sheets: Searched, Applied, Failed, Needs Human."""
from datetime import date, datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from app.config import ROOT
from app.db import session_scope
from app.models import Job, Application


def _header(ws, cols: list[str]):
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = font
        cell.fill = fill


def export_day(day: date | None = None) -> Path:
    day = day or date.today()
    start = datetime.combine(day, datetime.min.time())
    end = start + timedelta(days=1)

    out = ROOT / "logs" / "exports" / f"{day.isoformat()}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Searched
    ws = wb.active; ws.title = "Searched"
    _header(ws, ["discovered_at", "source", "company", "title", "location", "work_mode",
                 "ctc_min", "ctc_max", "url", "score", "status"])
    with session_scope() as db:
        rows = db.query(Job).filter(Job.discovered_at >= start, Job.discovered_at < end).all()
        for j in rows:
            ws.append([j.discovered_at.isoformat() if j.discovered_at else "", j.source, j.company, j.title,
                       j.location or "", j.work_mode or "", j.ctc_min, j.ctc_max,
                       j.url, j.score, j.status])

    # Applied
    ws = wb.create_sheet("Applied")
    _header(ws, ["applied_at", "source", "company", "title", "url", "outcome", "cover_letter_excerpt", "confirmation"])
    with session_scope() as db:
        rows = (db.query(Application).join(Job)
                .filter(Application.started_at >= start, Application.started_at < end,
                        Application.outcome.in_(["submitted", "dry_run"])).all())
        for a in rows:
            ws.append([a.started_at.isoformat(), a.job.source, a.job.company, a.job.title,
                       a.job.url, a.outcome, (a.cover_letter_md or "")[:200],
                       (a.confirmation_text or "")[:300]])

    # Failed
    ws = wb.create_sheet("Failed")
    _header(ws, ["attempted_at", "source", "company", "title", "url", "failure_reason", "screenshot_path"])
    with session_scope() as db:
        rows = (db.query(Application).join(Job)
                .filter(Application.started_at >= start, Application.started_at < end,
                        Application.outcome == "failed").all())
        for a in rows:
            ws.append([a.started_at.isoformat(), a.job.source, a.job.company, a.job.title,
                       a.job.url, a.failure_reason or "", a.screenshot_path or ""])

    # Needs human
    ws = wb.create_sheet("Needs Human")
    _header(ws, ["flagged_at", "source", "company", "title", "url", "reason"])
    with session_scope() as db:
        rows = (db.query(Application).join(Job)
                .filter(Application.started_at >= start, Application.started_at < end,
                        Application.outcome == "needs_human").all())
        for a in rows:
            ws.append([a.started_at.isoformat(), a.job.source, a.job.company, a.job.title,
                       a.job.url, a.failure_reason or ""])

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    wb.save(out)
    return out
